from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()

# Sheet 1: Kahvikupit
ws1 = wb.active
ws1.title = "1_Kahvikupit"

rows1 = [
    ["Tehtävä 1: Kahvikupit", "", ""],
    ["Oletukset", "", ""],
    ["Kappalehinta", 5, "eur/kpl"],
    ["Muuttuva kustannus / kuppi", 2, "eur/kpl"],
    ["Kiinteät kustannukset / vuosi", 10000, "eur"],
    ["", "", ""],
    ["Tämä vuosi", "", ""],
    ["Myyntimäärä", 5000, "kpl"],
    ["Liikevaihto", "=B3*B8", "hinta * määrä"],
    ["Muuttuvat kustannukset", "=B4*B8", "muuttuva kustannus * määrä"],
    ["Katetuotto", "=B9-B10", "liikevaihto - muuttuvat kustannukset"],
    ["Tulos", "=B11-B5", "katetuotto - kiinteät kustannukset"],
    ["", "", ""],
    ["Ensi vuosi (suunnitelma)", "", ""],
    ["Myyntimäärä", 6000, "kpl"],
    ["Liikevaihto", "=B3*B15", "hinta * määrä"],
    ["Muuttuvat kustannukset", "=B4*B15", "muuttuva kustannus * määrä"],
    ["Katetuotto", "=B16-B17", "liikevaihto - muuttuvat kustannukset"],
    ["Tulos", "=B18-B5", "katetuotto - kiinteät kustannukset"],
    ["", "", ""],
    ["Ensi vuoden vaihtoehto (6600 kpl, hinta 4,6)", "", ""],
    ["Kappalehinta", 4.6, "eur/kpl"],
    ["Myyntimäärä", 6600, "kpl"],
    ["Liikevaihto", "=B22*B23", "hinta * määrä"],
    ["Muuttuvat kustannukset", "=B4*B23", "muuttuva kustannus * määrä"],
    ["Katetuotto", "=B24-B25", "liikevaihto - muuttuvat kustannukset"],
    ["Kiinteät kustannukset", "=B5", "eur"],
    ["Tulos", "=B26-B27", "katetuotto - kiinteät kustannukset"],
    ["Katetuotto-%", "=B26/B24", "muotoile prosentiksi"],
]

for r in rows1:
    ws1.append(r)

# Sheet 2: Kuntosali
ws2 = wb.create_sheet("2_Kuntosali")
rows2 = [
    ["Tehtävä 2: Kuntosali", "", ""],
    ["Oletukset", "", ""],
    ["Asiakaskäynnit / kk", 1800, "kpl"],
    ["Hinta / käynti", 6, "eur"],
    ["Muuttuva kustannus / käynti", 1, "eur"],
    ["Kiinteät kustannukset / kk", 5000, "eur"],
    ["", "", ""],
    ["a) Katetuottolaskelma", "", ""],
    ["Liikevaihto", "=B3*B4", ""],
    ["Muuttuvat kustannukset", "=B3*B5", ""],
    ["Katetuotto", "=B9-B10", ""],
    ["Tulos", "=B11-B6", ""],
    ["", "", ""],
    ["b) Katetuottoprosentti", "=B11/B9", "muotoile prosentiksi"],
    ["", "", ""],
    ["c) Kriittinen piste euroina", "=B6/(1-B10/B9)", ""],
    ["", "", ""],
    ["d) Kriittinen piste asiakasmäärä", "=B16/B4", "kpl"],
    ["", "", ""],
    ["e) Varmuusmarginaali euroina", "=B9-B16", ""],
    ["Varmuusmarginaali-%", "=(B9-B16)/B9", "muotoile prosentiksi"],
    ["", "", ""],
    ["f) Tulos, jos asiakasmäärä laskee 15%", "", ""],
    ["Uusi asiakasmäärä", "=B3*(1-15%)", "kpl"],
    ["Uusi liikevaihto", "=B24*B4", ""],
    ["Uudet muuttuvat kustannukset", "=B24*B5", ""],
    ["Uusi katetuotto", "=B25-B26", ""],
    ["Uusi tulos", "=B27-B6", ""],
]
for r in rows2:
    ws2.append(r)

# Sheet 3: Matkatoimisto
ws3 = wb.create_sheet("3_Matkatoimisto")
rows3 = [
    ["Tehtävä 3: Matkatoimisto", "", ""],
    ["Oletukset", "", ""],
    ["Bussin ja kuljettajan vuokra", 2500, "eur/matka"],
    ["Oppaan palkkio", 1500, "eur/matka"],
    ["Hotelli + aamiainen / hlö", 130, "eur/hlö"],
    ["Bussin kapasiteetti", 50, "hlö"],
    ["Kiinteät kustannukset yhteensä", "=B3+B4", "eur"],
    ["", "", ""],
    ["a) Bussi täynnä, voittotavoite 2000 eur", "", ""],
    ["Matkustajamäärä", "=B6", "hlö"],
    ["Voittotavoite", 2000, "eur"],
    ["Tarvittava liikevaihto", "=B7+B5*B10+B11", ""],
    ["Hinta / hlö", "=B12/B10", "eur"],
    ["", "", ""],
    ["b) Käyttöaste 80%, voittotavoite 2000 eur", "", ""],
    ["Matkustajamäärä", "=B6*80%", "hlö"],
    ["Voittotavoite", 2000, "eur"],
    ["Tarvittava liikevaihto", "=B7+B5*B16+B17", ""],
    ["Hinta / hlö", "=B18/B16", "eur"],
    ["", "", ""],
    ["c) Bussi täynnä, voittotavoite 5000 eur", "", ""],
    ["Matkustajamäärä", "=B6", "hlö"],
    ["Voittotavoite", 5000, "eur"],
    ["Tarvittava liikevaihto", "=B7+B5*B22+B23", ""],
    ["Hinta / hlö", "=B24/B22", "eur"],
    ["", "", ""],
    ["d) Käyttöaste 50%, voittotavoite 1000 eur", "", ""],
    ["Matkustajamäärä", "=B6*50%", "hlö"],
    ["Voittotavoite", 1000, "eur"],
    ["Tarvittava liikevaihto", "=B7+B5*B28+B29", ""],
    ["Hinta / hlö", "=B30/B28", "eur"],
]
for r in rows3:
    ws3.append(r)

for ws in [ws1, ws2, ws3]:
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 34
    for c in ["A1", "A2", "A7", "A14", "A21", "A8", "A9", "A15", "A21", "A23", "A27"]:
        if ws[c].value:
            ws[c].font = Font(bold=True)

# Number formats
for ws in [ws1, ws2, ws3]:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
        cell = row[0]
        if isinstance(cell.value, str) and cell.value.startswith("="):
            pass

ws1["B29"].number_format = "0.00%"
ws2["B14"].number_format = "0.00%"
ws2["B21"].number_format = "0.00%"

output = "/home/joona/verkkosivu/tehtavat_katetuotto.xlsx"
wb.save(output)
print(output)
